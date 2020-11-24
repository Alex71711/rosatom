
from selenium import webdriver
from openpyxl import Workbook
import openpyxl, datetime, num2words, pymorphy2
from openpyxl.styles import Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

'''Открытие браузера, переход на страницу яндекса, нажатие на USD MOEX'''
driver = webdriver.Chrome('C:\chromedriver.exe')
driver.maximize_window()
driver.get('https://yandex.ru/')
driver.find_element_by_xpath('//*[@id="wd-_topnews"]/div/div[3]/div/div/div[1]/a').click()
driver.switch_to.window(driver.window_handles[1])


'''Формирование списков с данными для засение в эксель таблицу(строки 20-40)'''
kurs_dollar = driver.find_element_by_xpath('//*[@id="neo-page"]/div/div[2]/div[1]/div[1]/div[1]/div[2]/div/div[2]/div/div[2]').text.split('\n')
del kurs_dollar[0:3]

change_dollar = driver.page_source
change_dollar = change_dollar[change_dollar.find('"changes":')+11:change_dollar.find('"groups":')-2].split('},{')


driver.get('https://yandex.ru/')
driver.find_element_by_xpath('//*[@id="wd-_topnews"]/div/div[3]/div/div/div[2]/a').click()
driver.switch_to.window(driver.window_handles[2])


kurs_euro = driver.find_element_by_xpath('//*[@id="neo-page"]/div/div[2]/div[1]/div[1]/div[1]/div[2]/div/div[2]/div/div[2]').text.split('\n')
del kurs_euro[0:3]

change_euro = driver.page_source
change_euro = change_euro[change_euro.find('"changes":')+11:change_euro.find('"groups":')-2].split('},{')

for i in range(10):
    kurs_dollar[i*3+2]=change_dollar[i][change_dollar[i].find('"change":')+9:change_dollar[i].find('changePercentage')-2]
    kurs_euro[i*3+2]=change_euro[i][change_euro[i].find('"change":')+9:change_euro[i].find('changePercentage')-2]

driver.quit()

'''Создание эксель файла'''
wb = Workbook()
ws = wb.active

header = ['Дата_$','Курс_$','Изменени_$','Дата_E','Курс_E','Изменени_E','Курс_E/Курс_$']

ws.append(header)
'''Наполнение эксель файла данными(строки 52-88)'''
for i in range(10):
    for j in range(3):
        if j == 0:
            ws.cell(row=i + 2, column=j + 1).value = datetime.datetime.strptime(kurs_dollar[i * 3 + j],'%d.%m.%y').date()
        elif j == 1:
            ws.cell(row=i + 2, column=j + 1).value = float(kurs_dollar[i * 3 + j].replace(',', '.'))
            ws.cell(row=i + 2, column=j + 1).number_format = '#,##0.0000$'
        else:
            ws.cell(row=i + 2, column=j+1).value = float(kurs_dollar[i*3+j].replace(',','.'))
            ws.cell(row=i + 2, column=j+1).number_format = '#,##0.00'

    for l in range(3):
        if l == 0:
            ws.cell(row=i + 2, column=l + 4).value = datetime.datetime.strptime(kurs_euro[i * 3 + l],'%d.%m.%y').date()
        elif l == 1:
            ws.cell(row=i + 2, column=l + 4).value = float(kurs_euro[i * 3 + l].replace(',', '.'))
            ws.cell(row=i + 2, column=l + 4).number_format = '#,##0.0000€'
        else:
            ws.cell(row=i + 2, column=l+4).value = float(kurs_euro[i*3+l].replace(',','.'))
            ws.cell(row=i + 2, column=l+4).number_format = '#,##0.00'

    ws.cell(row=i + 2, column=7).value = '=E'+str(i + 2)+'/B'+str(i + 2)
    ws.cell(row=i + 2, column=7).number_format = '#,##0.00'

proverka = ['Автосумма:' , '=SUM(B2:B11)','=SUM(C2:C11)', ' ', '=SUM(E2:E11)', '=SUM(F2:F11)', '=SUM(G2:G11)' ]
ws.append(proverka)

'''Установка оптимальной ширины столбцов'''
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length*1.5

'''Выравние данных по центру ячейки'''
for row_cells in ws.iter_rows():
    for cell in row_cells:
        cell.alignment = Alignment(horizontal='center', vertical='center')

row_count = ws.max_row

wb.save("Kurs.xlsx")


'''Формирование сообщения в нужном склонении(строки 95-110)'''
morph = pymorphy2.MorphAnalyzer()
stroka = morph.parse('строка')[0]

stroka_str = stroka.make_agree_with_number(row_count).word

chislo_str = num2words.num2words(row_count, lang='ru')
chislo = morph.parse(chislo_str.split(' ')[-1])[0]
try:
    chislo = chislo.inflect({'femn'})
    chislo_str = chislo_str.split(' ')
    chislo_str[-1] = chislo.word
    chislo_str = ' '.join(chislo_str)
except:
    chislo_str = num2words.num2words(row_count, lang='ru')

text_massage = 'В эксель файле '+chislo_str+' '+ stroka_str

'''Функция отправки файла и сообщения на почту'''
def send_mail(send_from,send_to,text,files,password='',isTls=True):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = send_to
    msg['Date'] = formatdate(localtime = True)
    msg['Subject'] = 'Python email'
    msg.attach(MIMEText(text))

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(files, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename='+files)
    msg.attach(part)

    smtp = smtplib.SMTP('smtp.gmail.com',587)
    if isTls:
        smtp.starttls()
    smtp.login(send_from,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()


send_mail('send_from','send_to',text_massage, 'Kurs.xlsx', 'password')